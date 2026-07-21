import json
import openpyxl

wb = openpyxl.load_workbook('/sessions/blissful-stoic-babbage/mnt/uploads/営業管理.xlsx', data_only=True)

ws = wb['売上粗利表']
months = list(range(1, 13))
last_year_sales = [ws.cell(row=6, column=c).value or 0 for c in range(2, 14)]
this_year_target_sales = [ws.cell(row=7, column=c).value or 0 for c in range(2, 14)]
this_year_target_profit = [ws.cell(row=8, column=c).value or 0 for c in range(2, 14)]
necessary_expense = [ws.cell(row=11, column=c).value or 0 for c in range(2, 14)]

dash = wb['ダッシュボード']
profit_target_raw = dash.cell(row=4, column=7).value  # '\2,000,000'

settings = {
    "reps": ["新谷 壮央", "白木 寿樹", "川﨑 人志"],
    "aggregationYear": 2026,
    "monthlyProfitTarget": 2000000,
    "lastYearMonthlySales": last_year_sales,
    "monthlySalesTarget": this_year_target_sales,
    "monthlyProfitTargetByMonth": this_year_target_profit,
    "monthlyNecessaryExpense": necessary_expense,
}

with open('/sessions/blissful-stoic-babbage/mnt/outputs/app/seed_settings.json', 'w', encoding='utf-8') as f:
    json.dump(settings, f, ensure_ascii=False, indent=2)

print(json.dumps(settings, ensure_ascii=False, indent=2))
